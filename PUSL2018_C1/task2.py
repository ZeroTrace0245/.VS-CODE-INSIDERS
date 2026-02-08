#!/usr/bin/env python3
"""
dice_prob.py
Exact and simulated probabilities for sums of dice.

Usage examples:
  python task2.py --n_dice 10 --target 32 --trials 500 --out results/monte_carlo.xlsx
  python task2.py --n_dice 10 --target 32 --trials 500 --seed 123 --out results/monte_carlo.xlsx
"""
import argparse
import time
import random
import os
import statistics
from fractions import Fraction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Default output into results folder (matches task1)
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
DEFAULT_OUT = os.path.join(RESULTS_DIR, "monte_carlo.xlsx")


def exact_count(n_dice, target):
    """Compute exact number of sequences of n_dice dice (1..6) summing to target using DP (no numpy)."""
    if target < n_dice or target > 6 * n_dice:
        return 0
    max_sum = 6 * n_dice
    dp = [0] * (max_sum + 1)
    dp[0] = 1
    for _ in range(n_dice):
        ndp = [0] * (max_sum + 1)
        for s, v in enumerate(dp):
            if v:
                for face in range(1, 7):
                    ndp[s + face] += v
        dp = ndp
    return dp[target]


def simulate(n_dice, target, trials, rng):
    counts = []
    successes = 0
    for _ in range(trials):
        s = sum(rng.randint(1, 6) for _ in range(n_dice))
        counts.append(s)
        if s == target:
            successes += 1
    est_prob = successes / trials if trials > 0 else 0.0
    return successes, est_prob, counts


def load_or_create_workbook(path):
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception:
            backup = path + ".bak"
            try:
                os.replace(path, backup)
            except Exception:
                pass
    return Workbook()


def auto_size_columns(ws):
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 40)


def extract_monte_summary_rows(wb):
    sheet_name = "Monte Carlo Simulation"
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Summary":
            header_row = row_idx + 1
            headers = []
            col = 1
            while True:
                header_value = ws.cell(row=header_row, column=col).value
                if header_value is None:
                    break
                headers.append(header_value)
                col += 1
            data_rows = []
            current_row = header_row + 1
            while current_row <= ws.max_row and ws.cell(row=current_row, column=1).value is not None:
                entry = {}
                for index, header in enumerate(headers, start=1):
                    entry[header] = ws.cell(row=current_row, column=index).value
                data_rows.append(entry)
                current_row += 1
            parsed = []
            for raw in data_rows:
                parsed.append({
                    "N": int(raw["Number of Throws"]),
                    "mean_pi": float(raw["Mean π"]),
                    "std_pi": float(raw["Std Dev"]),
                    "mode_pi": raw["Mode π"],
                    "min_pi": float(raw["Min π"]),
                    "max_pi": float(raw["Max π"]),
                    "avg_time_s": float(raw["Average Time (s)"]),
                })
            return parsed
    return []


def summarize_monte_carlo(summary_rows):
    if not summary_rows:
        return []
    ordered = sorted(summary_rows, key=lambda item: item["N"])
    largest = ordered[-1]
    overall_min = min(item["min_pi"] for item in summary_rows)
    overall_max = max(item["max_pi"] for item in summary_rows)
    overall_mean = statistics.mean(item["mean_pi"] for item in summary_rows)
    return [
        ("Monte Carlo", f"Mean π @ N={largest['N']:,}", largest["mean_pi"]),
        ("Monte Carlo", f"Std Dev @ N={largest['N']:,}", largest["std_pi"]),
        ("Monte Carlo", "Overall min π estimate", overall_min),
        ("Monte Carlo", "Overall max π estimate", overall_max),
        ("Monte Carlo", "Average of mean π across N", overall_mean),
    ]


def summarize_dice_for_dashboard(exact_row, summary_row):
    return [
        ("Dice", "Exact probability (calculated)", round(exact_row["exact_probability"], 8)),
        ("Dice", "Exact probability (fraction)", exact_row["exact_probability_frac"]),
        ("Dice", f"Estimated probability (trials={summary_row['trials']})", round(summary_row["estimated_probability"], 8)),
        ("Dice", "Successful trials", summary_row["successes"]),
        ("Dice", "Simulation runtime (s)", round(summary_row["compute_time_s"], 6)),
        ("Dice", "Exact computation runtime (s)", round(exact_row["compute_time_s"], 6)),
    ]


def update_combined_sheet(wb, dice_rows):
    rows = []
    rows.extend(summarize_monte_carlo(extract_monte_summary_rows(wb)))
    rows.extend(dice_rows)
    sheet_name = "Combined Results"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Experiment", "Metric", "Value"])
    if rows:
        for row in rows:
            ws.append(row)
    else:
        ws.append(["-", "No results yet", "-"])
    auto_size_columns(ws)
    ws.freeze_panes = "A2"


def write_excel(out_path, exact_row, sim_counts, summary_row):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = load_or_create_workbook(out_path)
    sheet_name = "Dice Probability"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Trial", "Sum"])
    for index, total in enumerate(sim_counts, start=1):
        ws.append([index, total])
    ws.append([None])
    summary_title_row = ws.max_row + 1
    title_cell = ws.cell(row=summary_title_row, column=1, value="Summary")
    title_cell.font = Font(bold=True)
    ws.append(["Metric", "Value"])
    ws.append(["Exact probability (calculated)", round(exact_row["exact_probability"], 8)])
    ws.append(["Exact probability (fraction)", exact_row["exact_probability_frac"]])
    ws.append([f"Estimated probability (trials={summary_row['trials']})", round(summary_row["estimated_probability"], 8)])
    ws.append(["Successful trials", summary_row["successes"]])
    ws.append(["Total trials", summary_row["trials"]])
    ws.append(["Simulation runtime (s)", round(summary_row["compute_time_s"], 6)])
    ws.append(["Exact computation runtime (s)", round(exact_row["compute_time_s"], 6)])
    auto_size_columns(ws)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    update_combined_sheet(wb, summarize_dice_for_dashboard(exact_row, summary_row))
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Dice sum exact and simulated probabilities")
    parser.add_argument("--n_dice", type=int, default=10, help="Number of dice per trial")
    parser.add_argument("--target", type=int, default=32, help="Target sum to check")
    parser.add_argument("--trials", type=int, default=500, help="Number of simulation trials")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--out", type=str, default=DEFAULT_OUT, help="Excel output filename")
    args = parser.parse_args()

    rng = random.Random(args.seed)
    t0 = time.time()
    count = exact_count(args.n_dice, args.target)
    total = 6 ** args.n_dice
    exact_prob = Fraction(count, total) if total > 0 else Fraction(0, 1)
    t1 = time.time()

    successes, est_prob, counts = simulate(args.n_dice, args.target, args.trials, rng)
    t2 = time.time()

    exact_row = {
        "n_dice": args.n_dice,
        "target": args.target,
        "count": count,
        "total_outcomes": total,
        "exact_probability": float(exact_prob),
        "exact_probability_frac": f"{exact_prob.numerator}/{exact_prob.denominator}",
        "compute_time_s": t1 - t0,
    }
    summary_row = {
        "trials": args.trials,
        "successes": successes,
        "estimated_probability": est_prob,
        "compute_time_s": t2 - t1,
    }

    out_saved = write_excel(args.out, exact_row, counts, summary_row)

    print("Exact count and probability:")
    print(exact_row)
    print("\nSimulation summary:")
    print(summary_row)
    print(f"Saved results to: {out_saved}")


if __name__ == "__main__":
    main()
