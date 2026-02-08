#!/usr/bin/env python3
"""
dart_pi.py
Monte Carlo estimation of pi using random darts in [-1,1]x[-1,1].
Writes raw runs and summary to an Excel workbook (if available) and saves a PNG plot (if matplotlib is available).

Usage examples:
  python task1.py --out dart_results.xlsx
  python task1.py --Ns 1000 10000 100000 1000000 --repeats 10 --seed 123 --out my_dart.xlsx
  python task1.py --Ns 1000000 --repeats 5 --chunk_size 100000 --out dart_chunked.xlsx
"""
import argparse
import time
import math
import statistics
import random
import os
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Default output into results folder
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
DEFAULT_OUT = os.path.join(RESULTS_DIR, "monte_carlo.xlsx")


def run_one(rng, N, chunk_size=None):
    """Return (inside_count, pi_est). Implements chunking to reduce peak memory usage."""
    inside = 0
    if chunk_size is None or chunk_size >= N:
        for _ in range(N):
            x = rng.random() * 2.0 - 1.0
            y = rng.random() * 2.0 - 1.0
            if x * x + y * y <= 1.0:
                inside += 1
    else:
        remaining = N
        while remaining > 0:
            this = min(remaining, chunk_size)
            for _ in range(this):
                x = rng.random() * 2.0 - 1.0
                y = rng.random() * 2.0 - 1.0
                if x * x + y * y <= 1.0:
                    inside += 1
            remaining -= this
    return inside, 4.0 * inside / N


def compute_mode(values, round_digits=6):
    rounded = [round(v, round_digits) for v in values]
    most = Counter(rounded).most_common(1)
    return most[0][0] if most else None


def load_or_create_workbook(path):
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception:
            backup = path + ".bak"
            try:
                os.replace(path, backup)
            except Exception:
                pass
    wb = Workbook()
    return wb


def auto_size_columns(ws):
    for column in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is None:
                continue
            cell_length = len(str(cell.value))
            max_len = max(max_len, cell_length)
        ws.column_dimensions[column_letter].width = min(max_len + 2, 40)


def summarize_monte_carlo(summary_rows):
    if not summary_rows:
        return []
    ordered = sorted(summary_rows, key=lambda item: item["N"])
    largest = ordered[-1]
    overall_min = min(item["min_pi"] for item in summary_rows)
    overall_max = max(item["max_pi"] for item in summary_rows)
    overall_mean = statistics.mean(item["mean_pi"] for item in summary_rows)
    return [
        ("Monte Carlo", f"Mean π @ N={largest['N']:,}", largest["mean_pi"]),
        ("Monte Carlo", f"Std Dev @ N={largest['N']:,}", largest["std_pi"]),
        ("Monte Carlo", "Overall min π estimate", overall_min),
        ("Monte Carlo", "Overall max π estimate", overall_max),
        ("Monte Carlo", "Average of mean π across N", overall_mean),
    ]


def extract_dice_dashboard_rows(wb):
    sheet_name = "Dice Probability"
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    metrics = []
    for row_idx in range(1, ws.max_row + 1):
        marker = ws.cell(row=row_idx, column=1).value
        if marker == "Summary":
            header_row = row_idx + 1
            metric_col = 1
            value_col = 2
            current_row = header_row + 1
            while current_row <= ws.max_row:
                label = ws.cell(row=current_row, column=metric_col).value
                value = ws.cell(row=current_row, column=value_col).value
                if label is None:
                    break
                metrics.append(("Dice", label, value))
                current_row += 1
            break
    return metrics


def update_combined_sheet(wb, mc_rows):
    dashboard_rows = list(mc_rows)
    dashboard_rows.extend(extract_dice_dashboard_rows(wb))
    sheet_name = "Combined Results"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    headers = ["Experiment", "Metric", "Value"]
    ws.append(headers)
    if dashboard_rows:
        for entry in dashboard_rows:
            ws.append(entry)
    else:
        ws.append(["-", "No results yet", "-"])
    auto_size_columns(ws)
    ws.freeze_panes = "A2"


def write_excel(out_path, rows, summary_rows):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb = load_or_create_workbook(out_path)
    sheet_name = "Monte Carlo Simulation"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    headers = ["Number of Throws", "Run", "Inside Count", "π Estimate", "Time (s)"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["N"],
            row["repeat"],
            row["inside"],
            round(row["pi_est"], 8),
            round(row["time_s"], 6),
        ])
    ws.append([None])
    summary_title_row = ws.max_row + 1
    summary_title_cell = ws.cell(row=summary_title_row, column=1, value="Summary")
    summary_title_cell.font = Font(bold=True)
    summary_header_row = summary_title_row + 1
    summary_headers = [
        "Number of Throws",
        "Mean π",
        "Std Dev",
        "Mode π",
        "Min π",
        "Max π",
        "Average Time (s)",
        "Reference π",
    ]
    ws.append(summary_headers)
    for entry in summary_rows:
        ws.append([
            entry["N"],
            round(entry["mean_pi"], 8),
            round(entry["std_pi"], 8),
            entry["mode_pi"],
            round(entry["min_pi"], 8),
            round(entry["max_pi"], 8),
            round(entry["avg_time_s"], 6),
            round(math.pi, 8),
        ])
    auto_size_columns(ws)
    data_start = summary_header_row + 1
    data_end = ws.max_row
    if data_end >= data_start:
        chart = LineChart()
        chart.title = "Mean π convergence"
        chart.x_axis.title = "Number of Throws"
        chart.y_axis.title = "π estimate"
        chart.height = 12
        chart.width = 22
        categories = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        mean_series = Series(Reference(ws, min_col=2, min_row=data_start, max_row=data_end), title="Mean π")
        pi_series = Series(Reference(ws, min_col=8, min_row=data_start, max_row=data_end), title="π = 3.14159")
        chart.series.append(mean_series)
        chart.series.append(pi_series)
        chart.legend.position = "t"
        ws.add_chart(chart, f"J{summary_title_row}")
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    update_combined_sheet(wb, summarize_monte_carlo(summary_rows))
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Monte Carlo pi via darts")
    parser.add_argument("--Ns", nargs="+", type=int, default=[1000, 10000, 100000, 1000000],
                        help="List of N values (space separated)")
    parser.add_argument("--repeats", type=int, default=10, help="Repeats per N")
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--out", type=str, default=DEFAULT_OUT, help="Excel output filename")
    parser.add_argument("--chunk_size", type=int, default=None,
                        help="If set, generate points in chunks of this size (reduces peak memory)")
    args = parser.parse_args()

    rng = random.Random(args.seed)
    rows = []
    summary = []

    for N in args.Ns:
        estimates = []
        times = []
        for r in range(1, args.repeats + 1):
            t0 = time.time()
            inside, pi_est = run_one(rng, N, chunk_size=args.chunk_size)
            t1 = time.time()
            rows.append({"N": N, "repeat": r, "inside": inside, "pi_est": pi_est, "time_s": t1 - t0})
            estimates.append(pi_est)
            times.append(t1 - t0)
        mean = statistics.mean(estimates)
        std = statistics.stdev(estimates) if len(estimates) > 1 else 0.0
        mode_val = compute_mode(estimates, round_digits=6)
        summary.append({"N": N, "mean_pi": mean, "std_pi": std, "mode_pi": mode_val,
                        "min_pi": min(estimates), "max_pi": max(estimates), "avg_time_s": statistics.mean(times)})

    out_saved = write_excel(args.out, rows, summary)

    print(f"Saved results to: {out_saved}")
    print("Summary:")
    for s in summary:
        print(s)


if __name__ == "__main__":
    main()
